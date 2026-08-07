"""Bulk product import from CSV/XLS/XLSX price lists.

Column headers are matched case/accent-insensitively against a set of known
synonyms (``_HEADER_ALIASES``) so this works with human-edited spreadsheets
rather than requiring one exact template. See ``iter_product_rows`` for the
per-format readers and ``import_products_from_file`` for the upsert logic.

Design notes (see conversation / PR for the full rationale):
- Products are matched by ``sku`` when the row has one (idempotent re-import);
  rows without a SKU are always created as new products.
- Price/availability are only overwritten when the source row actually has a
  parseable price — re-importing a price-less catalog won't clobber pricing
  an admin set by hand afterwards. New products with no price import as
  unavailable with price 0.
- Category comes from (in order): an explicit "categoria" column, the xlsx
  sheet name, or ``default_category``.
- INGREDIENTES / DECLARACIÓN NUTRIMENTAL columns have no dedicated model
  field, so they're appended as labeled sections in ``description``.
- xlsx images are photos anchored to a cell (not URLs); they're extracted via
  openpyxl and saved straight to the ImageField, which runs them through the
  WebP compression pipeline (see ``shop/signals.py``) like any other upload.
"""

import csv
import unicodedata
from dataclasses import dataclass
from dataclasses import field as dataclass_field
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
import xlrd
from django.core.files.base import ContentFile
from django.db import transaction
from django.utils.text import slugify

from shop.models import Category, Product

_HEADER_ALIASES = {
    "sku": {"sku"},
    "producto": {"producto", "nombre", "nombre del producto"},
    "web_title": {"titulo de producto web", "titulo web", "titulo"},
    "net_content": {"contenido neto", "contenido", "presentacion"},
    "price": {"precio", "price"},
    "description": {"descripcion del producto", "descripcion", "description"},
    "ingredients": {"ingredientes", "ingredients"},
    "nutrition": {"declaracion nutrimental", "informacion nutrimental", "nutricion"},
    "category": {"categoria", "category"},
    "image": {"imagen", "image", "foto"},
}

_MIN_HEADER_MATCHES = 2


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    errors: list = dataclass_field(default_factory=list)  # [(row_label, message)]


def _normalize_header(value):
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.strip().lower().split())


def _clean_text(value):
    if value in (None, ""):
        return ""
    return str(value).strip()


def _build_field_map(headers):
    """Map canonical field name -> 0-indexed column position."""
    normalized = [_normalize_header(h) for h in headers]
    field_map = {}
    for canonical, aliases in _HEADER_ALIASES.items():
        for col_idx, header in enumerate(normalized):
            if header in aliases:
                field_map[canonical] = col_idx
                break
    return field_map


def _row_dict(values, field_map):
    return {name: values[idx] if idx < len(values) else None for name, idx in field_map.items()}


def _is_blank_row(row_dict):
    return not any(_clean_text(v) for v in row_dict.values())


def _parse_price(raw):
    if raw is None:
        return None
    if isinstance(raw, (int, float, Decimal)):
        return Decimal(str(raw))
    text = _clean_text(raw).replace("$", "").replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _build_product_name(row):
    producto = " ".join(_clean_text(row.get("producto")).split())
    net_content = " ".join(_clean_text(row.get("net_content")).split())
    if producto and net_content:
        return f"{producto} {net_content}"
    return producto or " ".join(_clean_text(row.get("web_title")).split())


def _build_description(row):
    parts = []
    description = _clean_text(row.get("description"))
    if description:
        parts.append(description)
    ingredients = _clean_text(row.get("ingredients"))
    if ingredients:
        parts.append(f"Ingredientes: {ingredients}")
    nutrition = _clean_text(row.get("nutrition"))
    if nutrition:
        parts.append(f"Declaración nutrimental:\n{nutrition}")
    return "\n\n".join(parts)


def _resolve_category_name(row, sheet_name, default_category):
    explicit = _clean_text(row.get("category"))
    if explicit:
        return explicit
    if sheet_name:
        return sheet_name.strip()
    if default_category:
        return default_category.strip()
    raise ValueError(
        "Sin categoría: no hay columna 'categoria', ni pestaña de hoja, ni --category."
    )


def _get_or_create_category(name):
    category = Category.objects.filter(name__iexact=name).first()
    if category:
        return category
    return Category.objects.create(name=name, slug=slugify(name))


def _resolve_local_image_path(base_dir, value):
    """Best-effort: treat a text 'imagen' cell (CSV/XLS) as a filesystem path."""
    text = _clean_text(value)
    if not text:
        return None
    candidate = Path(text)
    if not candidate.is_absolute():
        candidate = base_dir / candidate
    return candidate if candidate.is_file() else None


def _find_header_row(row_values_by_index, max_scan):
    """row_values_by_index(r) -> list of cell values for 1-indexed row r."""
    best_row, best_score = None, 0
    for r in range(1, max_scan + 1):
        values = row_values_by_index(r)
        if values is None:
            break
        score = len(_build_field_map(values))
        if score > best_score:
            best_row, best_score = r, score
    return best_row if best_score >= _MIN_HEADER_MATCHES else None


def _iter_csv_rows(path):
    raw = Path(path).read_bytes()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("cp1252")

    reader = csv.reader(text.splitlines())
    rows = list(reader)
    if not rows:
        return

    header_row_idx = _find_header_row(
        lambda r: rows[r - 1] if r <= len(rows) else None, min(10, len(rows))
    )
    if header_row_idx is None:
        return
    field_map = _build_field_map(rows[header_row_idx - 1])
    base_dir = Path(path).resolve().parent

    for i, raw_row in enumerate(rows[header_row_idx:], start=header_row_idx + 1):
        row = _row_dict(raw_row, field_map)
        if _is_blank_row(row):
            continue
        image_path = _resolve_local_image_path(base_dir, row.get("image"))
        image_bytes = image_path.read_bytes() if image_path else None
        image_ext = image_path.suffix.lstrip(".") if image_path else None
        yield None, i, row, (image_bytes, image_ext)


def _extract_row_images(worksheet, image_col_idx):
    """{0-indexed row -> raw bytes/format} for images anchored in ``image_col_idx``."""
    images = {}
    if image_col_idx is None:
        return images
    for img in getattr(worksheet, "_images", []):
        anchor = getattr(img, "anchor", None)
        frm = getattr(anchor, "_from", None)
        if frm is None or frm.col != image_col_idx:
            continue
        images[frm.row] = (img._data(), img.format)
    return images


def _iter_xlsx_rows(path, only_sheet=None):
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet_names = [only_sheet] if only_sheet else workbook.sheetnames

    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]

        def row_values(r, worksheet=worksheet):
            if r > worksheet.max_row:
                return None
            return [
                worksheet.cell(row=r, column=c).value for c in range(1, worksheet.max_column + 1)
            ]

        header_row_idx = _find_header_row(row_values, min(10, worksheet.max_row))
        if header_row_idx is None:
            continue
        field_map = _build_field_map(row_values(header_row_idx))
        images_by_row = _extract_row_images(worksheet, field_map.get("image"))

        for r in range(header_row_idx + 1, worksheet.max_row + 1):
            row = _row_dict(row_values(r), field_map)
            if _is_blank_row(row):
                continue
            image_info = images_by_row.get(r - 1, (None, None))
            yield sheet_name, r, row, image_info


def _iter_xls_rows(path, only_sheet=None):
    book = xlrd.open_workbook(path)
    sheet_names = [only_sheet] if only_sheet else book.sheet_names()

    for sheet_name in sheet_names:
        worksheet = book.sheet_by_name(sheet_name)

        def row_values(r, worksheet=worksheet):
            if r > worksheet.nrows:
                return None
            return worksheet.row_values(r - 1)

        header_row_idx = _find_header_row(row_values, min(10, worksheet.nrows))
        if header_row_idx is None:
            continue
        field_map = _build_field_map(row_values(header_row_idx))
        base_dir = Path(path).resolve().parent

        for r in range(header_row_idx + 1, worksheet.nrows + 1):
            row = _row_dict(row_values(r), field_map)
            if _is_blank_row(row):
                continue
            image_path = _resolve_local_image_path(base_dir, row.get("image"))
            image_bytes = image_path.read_bytes() if image_path else None
            image_ext = image_path.suffix.lstrip(".") if image_path else None
            yield sheet_name, r, row, (image_bytes, image_ext)


def iter_product_rows(file_path, sheet_name=None):
    """Yield ``(sheet_name, row_number, row_dict, (image_bytes, image_ext))``.

    ``row_dict`` keys are the canonical field names from ``_HEADER_ALIASES``
    (only the ones actually found in the file's header row are present).
    """
    ext = Path(file_path).suffix.lower().lstrip(".")
    if ext == "csv":
        yield from _iter_csv_rows(file_path)
    elif ext == "xlsx":
        yield from _iter_xlsx_rows(file_path, only_sheet=sheet_name)
    elif ext == "xls":
        yield from _iter_xls_rows(file_path, only_sheet=sheet_name)
    else:
        raise ValueError(f"Formato no soportado: .{ext} (usa .csv, .xls o .xlsx)")


def _upsert_product(sku, common_fields, price, image_content=None):
    """Create/update a Product. Price/availability only change when ``price``
    is not None, so re-imports don't clobber pricing set by hand in the admin.

    ``image_content`` (a ``ContentFile``, if given) is assigned as a plain
    attribute rather than via ``FieldFile.save()`` so it stays uncommitted
    until ``product.save()`` — that's what lets the pre_save WebP hook in
    shop/signals.py intercept and compress it, same as an admin upload.
    """
    product = Product.objects.filter(sku=sku).first() if sku else None
    created = product is None
    if created:
        product = Product(sku=sku)
        if price is None:
            common_fields = {**common_fields, "price": Decimal("0"), "available": False}
    if price is not None:
        common_fields = {**common_fields, "price": price, "available": True}

    for attr, value in common_fields.items():
        setattr(product, attr, value)
    if image_content is not None:
        product.image = image_content
    product.save()
    return product, created


def import_products_from_file(file_path, default_category=None, sheet_name=None, dry_run=False):
    """Import/update products from a CSV/XLS/XLSX price list.

    Returns an ``ImportResult`` with created/updated counts and a list of
    ``(row_label, message)`` for rows that failed and were skipped.
    """
    result = ImportResult()

    for sheet, row_num, row, (image_bytes, image_ext) in iter_product_rows(
        file_path, sheet_name=sheet_name
    ):
        row_label = f"{sheet}!{row_num}" if sheet else f"fila {row_num}"
        try:
            name = _build_product_name(row)
            if not name:
                raise ValueError("Falta el nombre del producto.")
            category_name = _resolve_category_name(row, sheet, default_category)
            sku = _clean_text(row.get("sku")) or None
            price = _parse_price(row.get("price"))
        except ValueError as exc:
            result.errors.append((row_label, str(exc)))
            continue

        if dry_run:
            exists = bool(sku and Product.objects.filter(sku=sku).exists())
            result.updated += 1 if exists else 0
            result.created += 0 if exists else 1
            continue

        try:
            with transaction.atomic():
                category = _get_or_create_category(category_name)
                common_fields = {
                    "name": name,
                    "slug": slugify(name)[:200],
                    "description": _build_description(row),
                    "category": category,
                }
                image_content = None
                if image_bytes:
                    ext = image_ext or "jpg"
                    image_content = ContentFile(
                        image_bytes, name=f"{sku or slugify(name)[:50]}.{ext}"
                    )
                product, created = _upsert_product(sku, common_fields, price, image_content)
        except Exception as exc:  # noqa: BLE001 - report and keep processing the rest
            result.errors.append((row_label, str(exc)))
            continue

        if created:
            result.created += 1
        else:
            result.updated += 1

    return result
